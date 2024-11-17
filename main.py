import pandas as pd
import openpyxl 
from openpyxl.styles import Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders 


#turning raw data from csv into excel sheet with accurate rows and data
file_path = './data.txt'
df = pd.read_csv(file_path)
excel_file_path = 'output.xlsx'
df.to_excel(excel_file_path, index=False, engine='openpyxl')
print(f"Excel file saved at: {excel_file_path}...")

#creating daily report sheet inside that workbook  
master_data = openpyxl.load_workbook("output.xlsx")
daily_sheet = master_data.create_sheet("daily_report")
print("created daily report...")
master_data.save(excel_file_path)

#populating dialy report sheet with headers
daily = master_data.worksheets[1]
daily_report_sheet_headers = ["name", "email", "id", "new total rewards"]
for i in range (len(daily_report_sheet_headers)):
    daily.cell(row=1, column=i+1, value=daily_report_sheet_headers[i])
master_data.save(excel_file_path)
print("populated daily report sheet...")

#populating names in daily sheet 
for row in range(2, master_data["Sheet1"].max_row + 1):
    first_name = master_data["Sheet1"][f"B{row}"].value
    last_name = master_data["Sheet1"][f"C{row}"].value
    full_name = f"{first_name} {last_name}"
    daily[f"A{row}"] = full_name
master_data.save(excel_file_path)
print("names updated...")

#populate emails 
for row in range(2, master_data["Sheet1"].max_row + 1):
    email = master_data["Sheet1"][f"D{row}"].value
    daily[f'B{row}'] = email
master_data.save(excel_file_path)
print("saved emails...")

#xlookup id nums
for row in range(2, master_data['Sheet1'].max_row + 1):
    daily[f'C{row}'] = f'=_xlfn.XLOOKUP(B{row}, Sheet1!D1:D101,Sheet1!A1:A101,,)'
master_data.save(excel_file_path)
print('id lookups saved...')

#populate total rewards sum 
for row in range(2, master_data['Sheet1'].max_row+1):
    rewards = master_data['Sheet1'][f'F{row}'].value 
    prevrewards = master_data['Sheet1'][f'G{row}'].value
    sum = rewards + prevrewards 
    master_data['daily_report'][f'D{row}'] = sum
master_data.save(excel_file_path)
print('formatted new rewards...')

#formatting daily report sheet
center_cells = Alignment(horizontal='center', vertical='center')  
for row in master_data['daily_report'].iter_rows():
    for cell in row:
        cell.alignment = center_cells
master_data.save(excel_file_path)
print('centered cells in report sheet...')


color_code = ColorScaleRule(start_type = 'num', start_value=0, start_color='FF9C9C',
mid_type = 'num', mid_value = 150, mid_color='89CFF0',
end_type='num', end_value = 500, end_color = '1DCB3A')
master_data['daily_report'].conditional_formatting.add('D2:D101', color_code)
print('color coded cells based on rewards amount...')
master_data.save(excel_file_path)


header_color = PatternFill(fill_type='solid', start_color='00ff00')
master_data['daily_report']['A1'].fill = header_color
master_data['daily_report']['B1'].fill = header_color
master_data['daily_report']['C1'].fill = header_color
master_data['daily_report']['D1'].fill = header_color
master_data.save(excel_file_path)
print('filling headers with header color...')


daily_report = master_data['daily_report']
for letter in ['A', 'B', 'C', 'D']:
    max_width = 0
    for row in range(1, daily_report.max_row+1):
        if len(str(daily_report[f'{letter}{row}'].value)) > max_width:
            max_width = len(str(daily_report[f'{letter}{row}'].value))
    daily_report.column_dimensions[letter].width = max_width + 1
master_data.save(excel_file_path)
print('adjusted width of cells for readability...')

#emailing report sheet to a test email 
secure_app_password =  'REPLACE THIS TEXT WITH APP SPECIFIC PASSWORD'
email_sender = 'lloydabonds@gmail.com'
email_reciever='johnsmithers206@gmail.com'

smtp_port = 587
smtp_server = 'smtp.gmail.com'

subject = 'Daily Report'

def send_emails(email_reciever):
    body = """
        line 1 
        line 2 
        sincerely, 
        lloyd 
     """ 
    
    msg = MIMEMultipart()
    msg['From'] = email_sender
    msg['To'] = email_reciever
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    filename = excel_file_path
    attachment = open(filename, 'rb')

    attachment_package = MIMEBase('application', 'octet-stream')
    attachment_package.set_payload((attachment).read())
    encoders.encode_base64(attachment_package)
    attachment_package.add_header('Content-Disposition', "attachment; filename=" + filename)
    msg.attach(attachment_package)

    text = msg.as_string()
    print('connecting to email server')
    TIE_server = smtplib.SMTP(smtp_server, smtp_port)
    TIE_server.starttls()
    TIE_server.login(email_sender, secure_app_password)
    print('connected to server')

    print('sending email...')
    TIE_server.sendmail(email_sender, email_reciever, text)
    print('email sent!')
    TIE_server.quit()

send_emails(email_reciever)